import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import os

class KPIApp:
    def __init__(self, root):
        self.root = root
        self.root.title("KPI Report Automator")
        self.root.geometry("600x400")
        self.root.configure(bg="#f0f0f0")

        self.style = ttk.Style()
        self.style.configure("TLabel", background="#f0f0f0", font=("Arial", 10))
        self.style.configure("TButton", font=("Arial", 10))
        self.style.configure("Header.TLabel", font=("Arial", 14, "bold"), background="#f0f0f0")

        self.raw_data_path = tk.StringVar()
        self.template_path = tk.StringVar()
        self.output_path = tk.StringVar()

        self.create_widgets()

    def create_widgets(self):
        # Header
        header = ttk.Label(self.root, text="KPI Report Generation Tool", style="Header.TLabel")
        header.pack(pady=20)

        # Main Frame
        frame = ttk.Frame(self.root, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        # Raw Data Section
        ttk.Label(frame, text="Raw Data File:").grid(row=0, column=0, sticky=tk.W, pady=10)
        ttk.Entry(frame, textvariable=self.raw_data_path, width=40, state="readonly").grid(row=0, column=1, padx=10)
        ttk.Button(frame, text="Browse", command=self.browse_raw).grid(row=0, column=2)

        # Template Section
        ttk.Label(frame, text="Template File:").grid(row=1, column=0, sticky=tk.W, pady=10)
        ttk.Entry(frame, textvariable=self.template_path, width=40, state="readonly").grid(row=1, column=1, padx=10)
        ttk.Button(frame, text="Browse", command=self.browse_template).grid(row=1, column=2)

        # Output Section
        ttk.Label(frame, text="Output File:").grid(row=2, column=0, sticky=tk.W, pady=10)
        ttk.Entry(frame, textvariable=self.output_path, width=40, state="readonly").grid(row=2, column=1, padx=10)
        ttk.Button(frame, text="Browse", command=self.browse_output).grid(row=2, column=2)

        # Generate Button
        self.generate_btn = ttk.Button(self.root, text="Generate Report", command=self.generate_report)
        self.generate_btn.pack(pady=20)

        # Status Label
        self.status_var = tk.StringVar()
        self.status_var.set("Ready")
        self.status_label = ttk.Label(self.root, textvariable=self.status_var, font=("Arial", 10, "italic"))
        self.status_label.pack(side=tk.BOTTOM, pady=10)

    def browse_raw(self):
        filename = filedialog.askopenfilename(title="Select Raw Data File", filetypes=[("Excel Files", "*.xlsx *.xls")])
        if filename:
            self.raw_data_path.set(filename)

    def browse_template(self):
        filename = filedialog.askopenfilename(title="Select Template File", filetypes=[("Excel Files", "*.xlsx *.xls")])
        if filename:
            self.template_path.set(filename)

    def browse_output(self):
        filename = filedialog.asksaveasfilename(title="Save Output As", defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if filename:
            self.output_path.set(filename)

    def generate_report(self):
        raw_file = self.raw_data_path.get()
        template_file = self.template_path.get()
        out_file = self.output_path.get()

        if not raw_file or not template_file or not out_file:
            messagebox.showerror("Error", "Please select all three file paths.")
            return

        self.status_var.set("Generating report... Please wait.")
        self.root.update()

        try:
            self.process_data(raw_file, template_file, out_file)
            self.status_var.set("Report generated successfully!")
            messagebox.showinfo("Success", f"Report successfully generated at:\n{out_file}")
        except Exception as e:
            self.status_var.set("Error generating report.")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

    def get_kpi_mapping(self):
        # Maps the exact clean KPI name derived from the Template to the exact Raw Data column name
        return {
            "ORA_3G_Cell Availability (%)": "ORA_3G_Cell Availability, excluding BLU (%)",
            "OCM_3G_CSSR_CS_CellPCH_URAPCH_New(%)": "ORA_3G_CSSR CS with Cell PCH/URA PCH (%)",
            "OCM_3G_CSSR_PS_CellPCH_URAPCH_New(%)": "ORA_3G_CSSR_PS_with_Cell_PCH (%)",
            "OCM_3G_Call_Drop_CS_New(%)": "ORA_3G_Drop Call Rate CS(%)",
            "ORA_3G_Call_Drop_PS_All_Data_Services_New(%)": "ORA_3G_Call Drop Data All Services(%)",
            "ORA_3G_CS Voice Traffic (Erl)": "ORA_3G_CS Voice Traffic (Erl)",
            "ORA_3G_Traffic Total Data_MAC (GB)": "ORA_3G_Traffic Total Data_MAC (GB)"
        }

    def process_data(self, raw_file, template_file, out_file):
        # 1. Load Template to get expected structure
        # The template seems to use Sheet2 for the main KPI layout
        xls_template = pd.ExcelFile(template_file)
        sheet_name = 'Sheet2' if 'Sheet2' in xls_template.sheet_names else xls_template.sheet_names[0]
        template_df = pd.read_excel(xls_template, sheet_name=sheet_name)

        # 2. Extract expected KPIs from template
        row_labels = template_df['Row Labels'].dropna().tolist()

        # We need to preserve order but uniqueness of KPIs since the template repeats them per site
        expected_kpis = []
        for label in row_labels:
            if label.startswith("Average of ") or label.startswith("Sum of "):
                if label not in expected_kpis:
                    expected_kpis.append(label)

        # 3. Load Raw Data
        xls_raw = pd.ExcelFile(raw_file)
        raw_sheet = 'Sheet0' if 'Sheet0' in xls_raw.sheet_names else xls_raw.sheet_names[0]
        raw_df = pd.read_excel(xls_raw, sheet_name=raw_sheet)

        # 4. Filter for last 10 hours
        # Convert Begin Time to datetime to sort and filter
        raw_df['Begin Time'] = pd.to_datetime(raw_df['Begin Time'])
        raw_df = raw_df.sort_values(by='Begin Time')

        # Get unique times and keep the last 10
        unique_times = raw_df['Begin Time'].dropna().unique()
        unique_times = pd.Series(unique_times).sort_values().tail(10).values

        # Filter raw_df to only include these top 10 hours
        filtered_raw = raw_df[raw_df['Begin Time'].isin(unique_times)]

        mapping = self.get_kpi_mapping()

        output_rows = []
        # Time headers
        time_headers = pd.Series(unique_times).dt.strftime('%Y-%m-%d %H:%M:%S').tolist()

        # Get list of Sites (NodeB Name) from raw data
        sites = filtered_raw['NodeB Name'].dropna().unique().tolist()

        for site in sites:
            site_data = filtered_raw[filtered_raw['NodeB Name'] == site]

            # Add Site Row
            output_rows.append({'Row Labels': site, **{t: None for t in time_headers}})

            for expected_kpi in expected_kpis:
                clean_kpi = expected_kpi.replace("Average of ", "").replace("Sum of ", "")

                raw_col = mapping.get(clean_kpi)
                if raw_col and raw_col in site_data.columns:
                    # Create a row dictionary
                    row_dict = {'Row Labels': expected_kpi}

                    for t_val, t_str in zip(unique_times, time_headers):
                        val_series = site_data[site_data['Begin Time'] == t_val][raw_col]
                        if not val_series.empty:
                            row_dict[t_str] = val_series.values[0]
                        else:
                            row_dict[t_str] = None

                    output_rows.append(row_dict)
                else:
                    # If mapping not found or col missing, insert empty row
                    row_dict = {'Row Labels': expected_kpi}
                    for t_str in time_headers:
                        row_dict[t_str] = None
                    output_rows.append(row_dict)

        # Create output dataframe
        final_df = pd.DataFrame(output_rows)

        # Write to Excel
        final_df.to_excel(out_file, index=False, sheet_name='Sheet1')

        # Apply openpyxl styling
        self.apply_excel_formatting(out_file, final_df, time_headers)

    def apply_excel_formatting(self, out_file, final_df, time_headers):
        wb = openpyxl.load_workbook(out_file)
        ws = wb['Sheet1']

        # Define Fills
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        # The user requested "rouge delave" (faded red) for bad values
        faded_red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Column indices for time headers
        # 'Row Labels' is column 1 (A), then time_headers are columns 2, 3, etc.
        num_cols = len(time_headers) + 1

        # Auto-fit column A width
        ws.column_dimensions['A'].width = 50
        for col_idx in range(2, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=num_cols), start=2):
            kpi_name = row[0].value
            if not kpi_name:
                continue

            # Formatting rules based on user prompt
            # - availability > or = 98.5 is green, else faded red
            # - CSSR (CS and PS) < 98.5 is faded red, else green (so >= 98.5 is green)
            # - Call drop (CS and PS) <= 0.7 is green, else faded red
            # - traffic is always white

            # Determine rule
            rule = None
            if "Availability" in str(kpi_name):
                rule = "availability"
            elif "CSSR" in str(kpi_name):
                rule = "cssr"
            elif "Call_Drop" in str(kpi_name) or "Drop Call Rate" in str(kpi_name) or "Call Drop Data" in str(kpi_name):
                rule = "call_drop"
            elif "Traffic" in str(kpi_name) or "Data_MAC" in str(kpi_name):
                rule = "traffic"

            for cell in row[1:]:
                val = cell.value
                if val is None or isinstance(val, str):
                    continue

                # Format to 2 decimal places if it's a number, but wait:
                # Raw data might be in decimal format for percentages (e.g. 1.0 instead of 100, 0.99 instead of 99%).
                # We need to format the numbers properly. Let's check magnitude to scale if needed.
                # Actually, looking at the template, availability and CSSR are percentages (0-100 scale).
                # Wait, if raw data gives 1.0 for 100%, we should multiply by 100 to match the template.

                # Check if it's a percentage kpi and the value is <= 1.5, it likely means it's decimal representation.
                # Or just check if the rule is availability/cssr/call_drop and value <= 1.0 (some call drops are tiny like 0.0004).

                # Actually, let's just multiply by 100 for percentage metrics if they are derived as decimals in raw data.
                # In the raw data, Cell Availability is 1.0, meaning 100%.
                # CSSR is 0.9982, meaning 99.82%.
                # Call Drop is 0.0004, meaning 0.04%.
                # The template expects these on a 100 scale.

                if rule in ["availability", "cssr", "call_drop"]:
                    # Assume values <= 1.5 are in decimal format and need *100
                    if val <= 1.5 and val >= -1.5:
                        val = val * 100
                        cell.value = val

                # Apply rules using the scaled values
                if rule == "availability":
                    if val >= 98.5:
                        cell.fill = green_fill
                    else:
                        cell.fill = faded_red_fill
                elif rule == "cssr":
                    if val >= 98.5:
                        cell.fill = green_fill
                    else:
                        cell.fill = faded_red_fill
                elif rule == "call_drop":
                    if val <= 0.7:
                        cell.fill = green_fill
                    else:
                        cell.fill = faded_red_fill
                elif rule == "traffic":
                    cell.fill = white_fill

                # Format to 2 decimal places if it's a number
                if isinstance(val, (int, float)):
                    cell.number_format = '0.00'

        wb.save(out_file)

if __name__ == "__main__":
    root = tk.Tk()
    app = KPIApp(root)
    root.mainloop()
