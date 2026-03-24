import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

class ReportGenerator:
    def __init__(self, template_path, output_path, shaped_data, times, kpis):
        self.template_path = template_path
        self.output_path = output_path
        self.shaped_data = shaped_data
        self.times = times
        self.kpis = kpis

    def _apply_formatting(self, cell, kpi_name, value):
        """Applies conditional formatting based on rules."""
        if value is None:
            return

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Rules exactly as defined:
        # Availability: >= 98.5 Green, else Pale Red
        if "Availability" in kpi_name:
            if value >= 98.5:
                cell.fill = green_fill
            else:
                cell.fill = red_fill

        # CSSR (CS & PS): < 98.5 Pale Red, else Green
        elif "CSSR" in kpi_name:
            if value < 98.5:
                cell.fill = red_fill
            else:
                cell.fill = green_fill

        # Call Drop (CS & PS): <= 0.7 Green, else Pale Red
        elif "Call_Drop" in kpi_name:
            if value <= 0.7:
                cell.fill = green_fill
            else:
                cell.fill = red_fill

        # Traffic CS: Always White (default) -> do nothing.

    def _set_cell_style(self, cell):
        """Basic border/alignment styling."""
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Calibri", size=11)

    def generate(self):
        # Create a new workbook with a single sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        # Headers setup
        ws.cell(row=1, column=1, value="Row Labels")
        for col_idx, t in enumerate(self.times, start=2):
            # Format time as 'YYYY-MM-DD HH:MM:SS'
            time_str = t.strftime('%Y-%m-%d %H:%M:%S')
            c = ws.cell(row=1, column=col_idx, value=time_str)
            self._set_cell_style(c)
            c.font = Font(bold=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
        ws.column_dimensions['A'].width = 60

        current_row = 2

        # Data rows
        for node in self.shaped_data:
            # Node header
            c = ws.cell(row=current_row, column=1, value=node)
            self._set_cell_style(c)
            c.font = Font(bold=True)
            current_row += 1

            # KPIs for node
            for kpi in self.kpis:
                c = ws.cell(row=current_row, column=1, value=kpi)
                self._set_cell_style(c)

                for col_idx, t in enumerate(self.times, start=2):
                    val = self.shaped_data[node][kpi].get(t)
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    self._set_cell_style(cell)

                    # Apply number format
                    if val is not None:
                        if "(%)" in kpi or "Rate" in kpi or "CSSR" in kpi or "Drop" in kpi:
                            cell.number_format = '0.00'
                        else:
                            cell.number_format = '#,##0.00'

                    # Conditional Formatting
                    self._apply_formatting(cell, kpi, val)

                current_row += 1

        wb.save(self.output_path)
        print(f"Report successfully generated at: {self.output_path}")

if __name__ == "__main__":
    from data_processor import DataProcessor
    from kpi_matcher import KPIMatcher

    raw_path = 'Performance Management-History Query-3G_KPI_Reporting_Template-DFBG6870-20260309082244.xlsx'
    template_path = 'Event_Performance Management-History Query-3G_KPI_Reporting_Template-DFBG6870-20251204091208.xlsx'
    out_path = 'test_output.xlsx'

    proc = DataProcessor(raw_path)
    df = proc.filter_last_10_hours()

    match = KPIMatcher(df)
    shaped, times, kpis = match.process()

    gen = ReportGenerator(template_path, out_path, shaped, times, kpis)
    gen.generate()
